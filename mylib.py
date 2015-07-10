# Windows XP
# create an environment variable PYTHONPATH and set it to the path to this file.
# This will allow python to reference this library file directly.
# re. Programing Python, O'reilly, Mark Lutz, p42 - 43

# courtesy of http://www.daniweb.com/code/snippet216610.html#
# time a function using time.time() and the a @ function decorator
# tested with Python24    vegaseat    21aug2005
from datetime import datetime
from pprint import pprint
from time import time
import xlwt


def time_this_function(func):
    def wrapper(*arg):
        t1 = time()
        res = func(*arg)
        t2 = time()
        print '%s took %0.3f ms' % (func.func_name, (t2-t1)*1000.0)
        return res
    return wrapper


class timing_device:
    def __init__(self, delay=3.0, activate=True):
        self.start = time()  # time zero
        self.delay = delay
        self.count = 0
        self.t = []
        self.activate = activate
        if self.activate:
            self.capture('Commence')

    def capture(self, comments):
        # previous event
        if not self.activate:
            return
        t = time() - self.start
        if t > self.delay:
            if self.t:
                # if there is already a first event
                # subtract the current start time with previous start time
                previous_event_duration = t-self.t[-1][1]
                self.t[-1][1] = previous_event_duration
                self.t.append([self.count, previous_event_duration, comments])
            else:
                # no previous event, this is the first event
                # just capture time of commencement of program till now
                previous_event_duration = t
                self.t.append([self.count, previous_event_duration, comments])
            self.count += 1

    def release(self):
        import csv
        if not self.activate:
            return
        header = ['S/no', 'Time taken', 'Event']

        fname = 'Event.csv'
        fhnd = open(fname, 'wb')
        csv_writer = csv.writer(fhnd)
        self.t.insert(0, header)
        csv_writer.writerows(self.t)
        fhnd.close()
        # for i in self.t:
        #     print "%d,%4.3f,%s" % (i[0], i[1], i[2])
        # self.t.insert(0, )
        del self.t
        self.t = []


def date_string():
    t = datetime.today()
    return "%04d%02d%02d" % (t.year, t.month, t.day)


def time_string():
    t = datetime.now()
    return "%02d%02d%02d" % (t.hour, t.minute, t.second)


def time_stringms():
    t = datetime.now()
    return "%02d%02d%02d%06d" % (t.hour, t.minute, t.second, t.microsecond)


def transpose(lists):
    """
    transpose a list from either row-oriented to column-oriented or vis-a-vis
    replace None values with defval
    courtesy of http://code.activestate.com/recipes/410687/

    note that graduation criteria file is column oriented, not the usual row-oriented
    1,2,3,4,5,6,7,...
    |,|,|,|,|,|,|,...
    |   | | | | |
    |   | | | |
    |   | |   |
        | |   |
        |     |

    so we'll need to sort of transpose the data from column oriented to row-oriented
    meaning something like
    1,-,-,-,-
    2,-
    3,-,-,-,-,-,-
    4,-,-,-,-,-
    5,-,-,-
    6,-,-,-,-,-,-
    7,-,-
    ...
    """
    if not lists:
        return []
    return map(lambda *row: list(row), *lists)


def transpose2(lists, defval=0):
    """
    replace None values with defval
    both function courtesy of http://code.activestate.com/recipes/410687/
    """
    if not lists:
        return []
    return map(lambda *row: [elem or defval for elem in row], *lists)


def ddmmmyyyyToDate(s):
    """
    gets a date string in the format 'dd-mmm-yyyy' and returns the datetime type
    """

    date = str(s).split('-')
    if len(date) != 3:
        dt = datetime.today()
    else:
        date_day, date_mth, date_year = date
        dt = datetime(int(date_year), monthStringToNum(date_mth), int(date_day))
    return dt


def monthStringToNum(s):
    """
    Provide the nth number given the name of the month, e.g. Jan -> 1; Dec -> 12
    """
    mth_dict = {'January': 1, 'February': 2, 'March': 3, 'April': 4, 'May': 5, 'June': 6,
                'July': 7, 'August': 8, 'September': 9, 'October': 10, 'November': 11, 'December': 12}
    # limiting the length so that 'Jan' can also be referenced
    index = s[:len(s)]
    return mth_dict[index] if index in mth_dict else ''


def average(seq=None):
    if seq is None:
        seq = []
    # print 'seq is', seq
    return sum(seq)/len(seq) if len(seq) else 0.0


def csString(s, sep=','):
    """
    compute Commas Separated String from list
    e.g. [1,2,3,4,5] -> "1,2,3,4,5"
    use string.split(s, delimiter) to reverse this function
    """
    return sep.join(s)


def commaSeparatedString(s, sep=','):
    if not s:
        return ''
    copyof_s = s[:]
    for i in range(len(s), 0, -1):
        if not copyof_s[i-1]:
            copyof_s.pop(i-1)
    return csString(copyof_s, sep)


def trim(s=''):
    return s.strip()


# return a string of list enclosed in brackets
def bracketString(s):
    return "" if s is None else "(%s)" % ','.join(["'%s'" % i.replace("'", "''") for i in s])


# return a string of list enclosed in brackets
def bbracketString(s):
    return "" if s is None else "(%s)" % ','.join(["\"%s\"" % i for i in s])


def chain(outer_s, outersep):
    """
    join a list of strings with a pre-defined symbol, e.g. +, will discard item(s)
    in the list if found to be empty; returns empty string if list is empty or
    contains empty strings only

    inserts a given delimiter between two non-zero strings, else just return
    either string without delimiter, return empty string if both string are empty
    """
    def cs(s, sep):
        dem = str(sep)
        s1 = str(s[0])
        s2 = str(s[1])
        return "%s %s %s" % (s1, dem, s2) if len(s1) and len(s2) else s1 + s2
    # recursive function, breaks a list into small list to be processed by
    # cs
    # e.g. [cs('a', 'b'), 'c', ...]
    #  =>  [cs('ab', 'c'), ...]
    #  =>  [cs('abc' .)..]

    def ct(s, sep):
        return ct([cs(s[:2], sep)] + s[2:], sep) if len(s) > 2 else cs(s, sep)
    return ct(outer_s, outersep) if isinstance(outer_s, list) else ""


def unistr(s):
    import unicodedata
    return s if s is None else unicodedata.normalize('NFKD', s).encode('ascii', 'ignore')


def ascii(s):
    """
    you can find out from http://www.asciitable.com/ on what is the entire
    ascii table, but this function only return the basic table
    truncate non ascii printable and extended ascii characters
    """
    s_value = ''
    if s:
        for c in s:
            # replace anything with the value of 0-31, >=127 with space
            s_value += c if 32 <= ord(c) < 127 else ' '
    return s_value


def replaceIllegal(s='', sep=''):
    """
    Remove illegal characters in a file name
    http://msdn.microsoft.com/en-us/library/Aa365247
    """
    illegal = '/\\:*?"<>|\'' + ''.join(chr(i) for i in range(32))
    if sep in illegal:
        sep = ''
    return "".join(sep if c in illegal else c for c in s)


def ReadSheetNames(fname):
    from os.path import exists
    import openpyxl
    if not exists(fname):
        return ()
    wb = openpyxl.load_workbook(filename=fname, read_only=True)
    result = wb.get_sheet_names()
    del wb
    return result


def ReadSpecificSheet(fname, shtName=''):
    """
    Given an excel spreadsheet file, Open the file using MS Excel and prompts
    user for a worksheet to select from get MS Excel to read the worksheets
    content and return to user as a python list
    """
    import openpyxl
    from os.path import exists
    from mylibwx import ChoiceDialog

    if not exists(fname):
        return ()
    wb = openpyxl.load_workbook(filename=fname, read_only=True)
    ws = wb.get_sheet_names()
    if shtName == '' or shtName not in ws:
        choice = ChoiceDialog("Select worksheet to extract data from:", fname, ws)
    else:
        choice = shtName
    if choice is None:
        return ()
    else:
        ws = wb['choice']
        used_range = [[c.value for c in row] for row in ws.rows]
        del wb
        return used_range


def xlsxShts(fname):
    """ Reads a given xlsx file (fname) and return a python list containing all the worksheet names
    """
    import os
    import openpyxl

    ffname = os.path.join(os.getcwd(), fname)
    set_name = os.path.exists(ffname)
    if set_name:
        wb = openpyxl.load_workbook(filename=ffname, read_only=True)
        sheets = wb.get_sheet_names()
    else:
        sheets = []
    # try to read the sheet specified
    return sheets


def WriteXls(fname, sht=1, data=None, append=False):
    """
    Writes a 2-D array (list) into a excel (xlsx) worksheet, if the sheetname is
    specified and already exists in the specified excel file name, the content
    in the worksheet is overwritten. Else a new sheet is created and the content
    written to the new worksheet.
    """
    import os
    import openpyxl
    from openpyxl.styles import Alignment
    from openpyxl.cell import get_column_letter

    if data is None:
        data = []
    ffname = os.path.join(os.getcwd(), fname)
    if os.path.exists(ffname):
        # wb = openpyxl.load_workbook(ffname, write_only=True)
        wb = openpyxl.load_workbook(filename=ffname)
    else:
        # wb = openpyxl.Workbook(write_only=True)
        wb = openpyxl.Workbook()

    # print "wb.read_only is %s" % ("True" if wb.read_only else "False")
    # print "wb.write_only is %s" % ("True" if wb.write_only else "False")
    if isinstance(sht, str):
        sheet_name = sht    # otherwise, just use the name given
    else:                   # if 1,2,3 ... is given, default to 'Sheet1', 'Sheet2', ...
        sheet_name = "Sheet%s" % sht

    if sheet_name in wb.get_sheet_names():
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)
    _cell = ws.cell
    for i, line in enumerate(data):
        if line is None:
            continue
        try:
            for j, cell in enumerate(line):
                try:
                    if cell is not None:
                        # _cell(i+1, j+1).NumberFormat = '@'
                        # inputs by Charlie Clark, @https://groups.google.com/forum/#!topic/openpyxl-users/hdTAidKQVpE
                        _cell(row=i+1, column=j+1).alignment = Alignment(wrap_text=False)
                        _cell(row=i+1, column=j+1).value = cell
                except:
                    pprint(cell)
                    raise
        except:
            pprint(line)
            raise
    for i, col in enumerate(ws.columns):
        column_list = []
        for c in col:
            if c.value is None:
                continue
            elif type(c.value) in (int, float):
                column_list.append(len(str(c.value)))
            else:
                column_list.append(len(c.value))
        ws.column_dimensions[get_column_letter(i+1)].width = max(column_list)+5

    wb.save(ffname)


def RefreshXlsx(fname):
    from os.path import exists
    from win32com.client import Dispatch

    if not exists(fname):
        return ()
    xlapp = Dispatch('Excel.Application')
    # xlapp = EnsureDispatch('Excel.Application')

    xlapp.DisplayAlerts = False
    xlapp.Visible = True
    xlapp.ScreenUpdating = True

    xlBook = xlapp.Workbooks.Open(fname)
    xlBook.RefreshAll()

    xlBook.Save()
    xlBook.Close()
    xlapp.Quit()
    del xlapp


def ReadxlsToList(fname, sht=1):
    """
    Uses PyWin32, directs Excel to open a xlsx spreadsheet, reads a specific
    worksheet and return its content in a python list
    Re-written using openpyxl - 14 May 2015
    """
    from os.path import exists
    import openpyxl

    if not exists(fname):
        return ()
    wb = openpyxl.load_workbook(filename=fname, read_only=True)
    ws = wb[sht]
    # dump = xlSheet.UsedRange

    used_range = [[c.value for c in row] for row in ws.rows]

    del wb
    # r_value = transpose(dump)
    return used_range


def save2xlsx(fname, output_dump=None):
    """
    Do a simple : create empty xlsx, paste python list as content, save as given
    fname, close and quit MS Excel
    """

    from win32com.client import Dispatch
    if output_dump is None:
        output_dump = []
    xlapp = Dispatch("excel.application")
    xlapp.DisplayAlerts = False
    xlapp.Visible = True
    xlapp.ScreenUpdating = True

    # create a new spreadsheet
    xlbook = xlapp.Workbooks.Add()

    # use the first worksheet
    sht1 = xlbook.Sheets("Sheet1")
    # inserts all the accepted claims
    address = list2exceladdress(output_dump)
    pprint(address)
    ur1 = sht1.Range(address)
    ur1.Value = output_dump
    xlbook.SaveAs(fname)    # save the spreadsheet
    xlbook.Close()              # close the workbook
    xlapp.DisplayAlerts = True
    xlapp.Quit()                # quit Excel


def rearrange(ilist=None, width=1):
    """
        Breaks up a long list into 2-dimensional list, with n * width (number of items in each list)
        e.g. split_list(range(10), 3) gives [[0, 1, 2], [3, 4, 5], [6, 7, 8], [9]]
    """
    if ilist is None:
        ilist = []
    if width < 1:
        return []
    # reduced lines of code, courtesy of http://www.how2code.co.uk/2013/04/how-to-split-a-list-into-chunks-in-python/
    return [ilist[x:x+width] for x in range(0, len(ilist), width)]


def rearrangev(ilist=None, width=None):
    """
        Breaks up a long list into 2-dimensional list, with specific size list given by a list of numbers
        e.g. split_list([0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19], [3,7,4,7,5])
        gives [
        [0, 1, 2],
        [3, 4, 5, 6, 7, 8, 9],
        [10, 11, 12, 13],
        [14, 15, 16, 17, 18, 19, 20],
        [21, 22, 23, 24, 25],
        [26, 27, 28, 29]
        ]
    """
    if ilist is None:
        ilist = []
    if width is None:
        width = []
    if len(width) < 1:
        return []
    size = len(ilist)
    ps = [sum(j for j in width[:i]) for i in xrange(len(width)+1)] + ([size] if size > sum(width) else [])
    return [ilist[start:end] for start, end in zip(ps[:-1], ps[1:])]


def split_list(ilist=None, divisor=1):
    """
        Breaks up a long list into 2-dimensional list, with divisor * n number of items in each list
        e.g. split_list(range(10), 3) gives [[0, 1, 2, 3], [4, 5, 6], [7, 8, 9]]
    """
    if ilist is None:
        ilist = []
    if divisor < 1:
        return []
    rlist = []
    size = len(ilist)
    remainder = size % divisor
    width = size // divisor
    start = 0
    while start < size:
        w = start + width
        if remainder > 0:
            remainder -= 1
            w += 1
        if w > size:
            w = size
        rlist.append(ilist[start:w])
        start = w
    return rlist


def atomic_copy(ilist=None, start=0, end=0):
    if ilist is None:
        ilist = []
    # t = type(ilist).__name__
    l = len(ilist)
    start = l - (start % l)
    end = l - (end % l)
    s = min([start, end])
    e = max([start, end])
    return ilist[s:e]


def is_link(lnk='', list_of_link_extension=None):
    import os
    if list_of_link_extension is None:
        list_of_link_extension = []
    if lnk is None:
        return False
    url_link, ext = os.path.splitext(lnk.lower())
    # assuming no extension and starts with 'http://' as a link
    try:
        # ext is in the form '.jpg' or '.html', so use ext[1:] to rid of the starting '.'
        val = ext in list_of_link_extension and url_link.startswith('http://')
    # print "%s : %s" % (str(val), lnk)
    except:
        val = False
    return val


def is_http_link(lnk=''):
    possible_http_extension = [".html", ".htm", ".php", ".shtml", ".mhtml", ".asp", ".aspx"]
    return is_link(lnk, possible_http_extension)


def is_pic_link(lnk=''):
    pic_types = ['jpg', 'jpeg', 'png', 'tiff', 'gif', 'bmp', 'pcx', 'ppm', 'jpeg']
    return is_link(lnk, pic_types)


def init_br():
    """
    http://stockrt.github.com/p/emulating-a-browser-in-python-with-mechanize/
     Browser
    """
    import mechanize
    import cookielib
    br = mechanize.Browser()

    # Cookie Jar
    cj = cookielib.LWPCookieJar()
    br.set_cookiejar(cj)

    # Browser options
    br.set_handle_equiv(True)
    # br.set_handle_gzip(True)
    br.set_handle_redirect(True)
    br.set_handle_referer(True)
    br.set_handle_robots(False)

    # Follows refresh 0 but not hangs on refresh > 0
    br.set_handle_refresh(mechanize._http.HTTPRefreshProcessor(), max_time=1)

    # Want debugging messages?
    # br.set_debug_http(True)
    # br.set_debug_redirects(True)
    # br.set_debug_responses(True)

    # User-Agent (this is cheating, ok?)
    br.addheaders = [('User-agent', 'Mozilla/4.0 (compatible; MSIE 8.0; Windows NT 6.1)')]
    return br


def two_dim_list2xls(sht=xlwt.Worksheet, data=None, col_width=None, x_offset=0, y_offset=0):
    from decimal import Decimal
    if data is None:
        data = []
    if col_width is None:
        col_width = []

    for y, line in enumerate(data):
        for x, cell in enumerate(line):
            try:
                if isinstance(cell, unicode):
                    tcell = cell.encode('utf-8')
                elif isinstance(cell, Decimal) or isinstance(cell, float):
                    tcell = float(cell)
                elif isinstance(cell, int):
                    tcell = int(cell)
                elif type(cell) in [list, dict, str]:
                    tcell = "%s" % cell
                else:
                    tcell = cell
                sht.write(y + y_offset, x + x_offset, tcell)
            except:
                pass    # None type, just ignore
    for i, width in enumerate(col_width):
        sht.col(i + y_offset).width = int(sht.col(i + y_offset).width * width)


def list2exceladdress(lst=None):
    if lst is None:
        lst = []
    last_cell = ""
    from string import ascii_uppercase
    try:
        maxX = max(len(row) for row in lst)
    except:
        raise
    Y = len(lst)
    llen = len(ascii_uppercase)
    while maxX > llen:
        last_cell = ascii_uppercase[(maxX % llen)-1] + last_cell
        maxX /= llen
    last_cell = ascii_uppercase[(maxX % llen)-1] + last_cell + str(Y)
    return "A1:%s" % last_cell


def ptdt(pt):
    """
    Convert PyTime to datetime
    """
    try:
        return datetime(
            year=pt.year,
            month=pt.month,
            day=pt.day,
            hour=pt.hour,
            minute=pt.minute,
            second=pt.second)
    except:
        return datetime(1900, 1, 1, 0, 0, 0)


if __name__ == '__main__':
    # print "'%s'" % chain(['',''],'+')
    # print csString([1,2,3])
    # raw_input('Press any key to continue...')
    # WriteXls('test.xlsx', 4, [map(str, range(i)) for i in range(1,10)])
    WriteXls('test.xlsx', 4, map(str, range(5)))
    # D:\Documents\Doc\DIDM\AY1011Sem2StudentSelectionsData(21Aug2010).xls
    pprint(rearrange(range(20), 3))
    pprint(rearrangev(range(30), [3, 7, 4, 7, 5]))
    pass
